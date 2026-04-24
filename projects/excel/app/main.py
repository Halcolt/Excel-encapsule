from flask import Flask, render_template, request, redirect, url_for, flash, session, g, send_file
import pandas as pd
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import tempfile
from pathlib import Path
import uuid
import json
from io import BytesIO
from io import StringIO
import re
from urllib.parse import urlparse
import os
import logging
import threading
import time
import csv
import secrets
from collections import Counter
from datetime import date, datetime, timedelta

app = Flask(__name__)

# ----------
# App config
# ----------
def _get_env_int(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, default))
    except Exception:
        return default


def _configure_app(a: Flask) -> None:
    a.secret_key = os.getenv("FLASK_SECRET_KEY", "dev")
    max_mb = _get_env_int("MAX_UPLOAD_MB", 16)
    a.config["MAX_CONTENT_LENGTH"] = max_mb * 1024 * 1024


_configure_app(app)

# Temp upload root (system temp dir by default)
UPLOAD_ROOT = Path(os.getenv("UPLOAD_ROOT", str(Path(tempfile.gettempdir()) / "excel_viewer_uploads")))
UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)

# TTL for uploaded sessions (hours)
UPLOAD_TTL_HOURS = _get_env_int("UPLOAD_TTL_HOURS", 24)

# Logger
logger = logging.getLogger("excel_viewer")
if not logger.handlers:
    logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO"))

ALLOWED_EXTENSIONS = {"xlsx", "csv"}
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1258", "cp1252", "latin1")
CSV_DELIMITERS = [",", ";", "\t", "|"]
COLUMN_FORMAT_PRESETS = (
    {"id": "original", "label_key": "format_original"},
    {"id": "text", "label_key": "format_text"},
    {"id": "general", "label_key": "format_general"},
    {"id": "integer", "label_key": "format_integer"},
    {"id": "decimal_2", "label_key": "format_decimal_2"},
    {"id": "percent_2", "label_key": "format_percent_2"},
    {"id": "date_dmy", "label_key": "format_date_dmy"},
    {"id": "date_mdy", "label_key": "format_date_mdy"},
    {"id": "date_ymd", "label_key": "format_date_ymd"},
    {"id": "datetime_dmy_hm", "label_key": "format_datetime_dmy_hm"},
)
PRESET_NUMBER_FORMATS = {
    "text": "@",
    "general": "General",
    "integer": "0",
    "decimal_2": "0.00",
    "percent_2": "0.00%",
    "date_dmy": "dd-mm-yyyy",
    "date_mdy": "mm-dd-yyyy",
    "date_ymd": "yyyy-mm-dd",
    "datetime_dmy_hm": "dd-mm-yyyy hh:mm",
}


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def get_csrf_token() -> str:
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token


def _request_csrf_token() -> str:
    token = request.form.get("csrf_token") or request.headers.get("X-CSRFToken")
    if not token and request.is_json:
        payload = request.get_json(silent=True) or {}
        token = payload.get("csrf_token")
    return token or ""


@app.before_request
def _validate_csrf_token():
    if request.method != "POST":
        return None
    expected = session.get("_csrf_token")
    supplied = _request_csrf_token()
    if not expected or not supplied or not secrets.compare_digest(expected, supplied):
        return {"error": tr("flash_invalid_csrf")}, 400
    return None


def get_token_dir(token: str) -> Path:
    d = UPLOAD_ROOT / token
    d.mkdir(parents=True, exist_ok=True)
    return d


def _sniff_csv_delimiter(text: str) -> str | None:
    sample = text[:65536]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=CSV_DELIMITERS)
        return dialect.delimiter
    except csv.Error:
        return None


def read_csv_smart(path: Path) -> tuple[pd.DataFrame, dict[str, str]]:
    raw = path.read_bytes()
    last_error: Exception | None = None

    for encoding in CSV_ENCODINGS:
        try:
            text = raw.decode(encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
            continue

        detected = _sniff_csv_delimiter(text)
        candidates = [detected] if detected else []
        candidates.extend(d for d in CSV_DELIMITERS if d not in candidates)
        parsed: list[tuple[pd.DataFrame, dict[str, str]]] = []

        for delimiter in candidates:
            try:
                df = pd.read_csv(StringIO(text), sep=delimiter)
                parsed.append((df, {"encoding": encoding, "delimiter": delimiter}))
            except Exception as exc:
                last_error = exc

        if parsed:
            if detected:
                return parsed[0]
            return max(parsed, key=lambda item: len(item[0].columns))

    raise last_error or ValueError("Could not read CSV file")


def _normalize_cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value)


def _looks_like_datetime_format(fmt: str) -> bool:
    fmt_lower = (fmt or "").lower()
    return any(token in fmt_lower for token in ("h", "s", "am/pm"))


def _looks_like_date_format(fmt: str) -> bool:
    fmt_lower = (fmt or "").lower()
    return any(token in fmt_lower for token in ("d", "m", "y")) and not any(
        token in fmt_lower for token in ("0.00%", "#,##0", "@")
    )


def _infer_excel_cell_type(cell) -> str:
    value = cell.value
    fmt = cell.number_format or ""
    if value is None:
        return ""
    if cell.is_date or isinstance(value, datetime):
        return "datetime" if _looks_like_datetime_format(fmt) else "date"
    if isinstance(value, date):
        return "date"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        fmt_lower = fmt.lower()
        if "%" in fmt:
            return "percent"
        if "[$" in fmt_lower or any(symbol in fmt for symbol in ("$", "€", "£", "¥", "₫")):
            return "currency"
        if isinstance(value, int) or (isinstance(value, float) and float(value).is_integer()):
            return "integer"
        return "decimal"
    return "text"


def _infer_series_type(series: pd.Series) -> str:
    if series.empty:
        return "text"
    dtype = str(series.dtype).lower()
    if "datetime" in dtype:
        return "datetime"
    if "int" in dtype:
        return "integer"
    if "float" in dtype or "double" in dtype:
        return "decimal"
    if "bool" in dtype:
        return "boolean"
    return "text"


def _default_format_for_type(source_type: str) -> str:
    return {
        "date": "dd-mm-yyyy",
        "datetime": "dd-mm-yyyy hh:mm",
        "integer": "0",
        "decimal": "0.00",
        "percent": "0.00%",
        "currency": "#,##0.00",
        "text": "@",
    }.get(source_type, "General")


def _build_csv_column_metadata(df: pd.DataFrame) -> list[dict]:
    metadata = []
    for column_name in df.columns:
        series = df[column_name] if column_name in df else pd.Series(dtype="object")
        source_type = _infer_series_type(series)
        metadata.append(
            {
                "header": _normalize_cell_text(column_name),
                "source_type": source_type,
                "original_number_format": _default_format_for_type(source_type),
                "selected_preset": "original",
            }
        )
    return metadata


def _build_excel_column_metadata(path: Path, sheet_name: str, df: pd.DataFrame) -> list[dict]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        sheet = workbook[sheet_name]
        metadata: list[dict] = []
        for col_idx, column_name in enumerate(df.columns, start=1):
            source_types: list[str] = []
            number_formats: list[str] = []
            for row_idx in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                cell_type = _infer_excel_cell_type(cell)
                if cell_type:
                    source_types.append(cell_type)
                fmt = cell.number_format or ""
                if fmt and fmt != "General":
                    number_formats.append(fmt)
            source_type = Counter(source_types).most_common(1)[0][0] if source_types else _infer_series_type(df.iloc[:, col_idx - 1])
            original_number_format = (
                Counter(number_formats).most_common(1)[0][0]
                if number_formats
                else _default_format_for_type(source_type)
            )
            metadata.append(
                {
                    "header": _normalize_cell_text(column_name),
                    "source_type": source_type,
                    "original_number_format": original_number_format,
                    "selected_preset": "original",
                }
            )
        return metadata
    finally:
        workbook.close()


def _load_sheet_dataframe(path: Path, filename: str, sheet_name: str | None = None) -> tuple[pd.DataFrame, list[dict], str | None]:
    ext = filename.rsplit(".", 1)[1].lower()
    if ext == "csv":
        df, _ = read_csv_smart(path)
        df = df.fillna("")
        return df, _build_csv_column_metadata(df), None

    xls = pd.ExcelFile(path)
    target_sheet = sheet_name or (xls.sheet_names[0] if xls.sheet_names else None)
    if not target_sheet:
        raise ValueError("No sheets found in the workbook.")
    df = xls.parse(target_sheet).fillna("")
    column_metadata = _build_excel_column_metadata(path, target_sheet, df)
    return df, column_metadata, target_sheet


def _parse_numeric_value(raw_value, allow_percent: bool = False):
    text = _normalize_cell_text(raw_value).strip()
    if not text:
        return None
    normalized = text.replace(",", "").replace(" ", "")
    is_percent = normalized.endswith("%")
    if is_percent:
        normalized = normalized[:-1]
    try:
        number = float(normalized)
    except ValueError:
        return None
    if allow_percent:
        if is_percent or abs(number) > 1:
            return number / 100
        return number
    if number.is_integer():
        return int(number)
    return number


def _parse_datetime_value(raw_value, prefer_dayfirst: bool) -> datetime | None:
    text = _normalize_cell_text(raw_value).strip()
    if not text:
        return None
    explicit_formats = ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m-%d-%Y", "%m/%d/%Y"]
    explicit_datetimes = [
        "%d-%m-%Y %H:%M",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d %H:%M",
        "%m-%d-%Y %H:%M",
        "%m/%d/%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%m-%d-%Y %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
    ]
    ordered_formats = (explicit_datetimes + explicit_formats) if prefer_dayfirst else (
        explicit_datetimes[2:] + explicit_datetimes[:2] + explicit_formats[2:] + explicit_formats[:2]
    )
    for fmt in ordered_formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    parsed = pd.to_datetime(text, errors="coerce", dayfirst=prefer_dayfirst)
    if pd.isna(parsed):
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=not prefer_dayfirst)
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def _coerce_export_cell(raw_value, column_meta: dict) -> tuple[object, str]:
    text = _normalize_cell_text(raw_value)
    trimmed = text.strip()
    preset = (column_meta or {}).get("selected_preset") or "original"
    source_type = (column_meta or {}).get("source_type") or "text"
    original_number_format = (column_meta or {}).get("original_number_format") or "General"

    if trimmed == "":
        return "", PRESET_NUMBER_FORMATS.get(preset, original_number_format)

    if preset == "text":
        return text, PRESET_NUMBER_FORMATS["text"]

    if preset == "general":
        parsed_number = _parse_numeric_value(trimmed)
        return (parsed_number, "General") if parsed_number is not None else (text, "General")

    if preset in {"integer", "decimal_2"}:
        parsed_number = _parse_numeric_value(trimmed)
        if parsed_number is None:
            return text, PRESET_NUMBER_FORMATS[preset]
        if preset == "integer":
            parsed_number = int(round(float(parsed_number)))
        return parsed_number, PRESET_NUMBER_FORMATS[preset]

    if preset == "percent_2":
        parsed_number = _parse_numeric_value(trimmed, allow_percent=True)
        return (parsed_number, PRESET_NUMBER_FORMATS[preset]) if parsed_number is not None else (text, PRESET_NUMBER_FORMATS[preset])

    if preset in {"date_dmy", "date_mdy", "date_ymd", "datetime_dmy_hm"}:
        prefer_dayfirst = preset in {"date_dmy", "datetime_dmy_hm"}
        parsed_dt = _parse_datetime_value(trimmed, prefer_dayfirst=prefer_dayfirst)
        if parsed_dt is None:
            return text, PRESET_NUMBER_FORMATS[preset]
        if preset == "datetime_dmy_hm":
            return parsed_dt, PRESET_NUMBER_FORMATS[preset]
        return parsed_dt.date(), PRESET_NUMBER_FORMATS[preset]

    if source_type in {"integer", "decimal", "currency"}:
        parsed_number = _parse_numeric_value(trimmed)
        return (parsed_number, original_number_format) if parsed_number is not None else (text, original_number_format)

    if source_type == "percent":
        parsed_number = _parse_numeric_value(trimmed, allow_percent=True)
        return (parsed_number, original_number_format) if parsed_number is not None else (text, original_number_format)

    if source_type in {"date", "datetime"}:
        parsed_dt = _parse_datetime_value(trimmed, prefer_dayfirst=True)
        if parsed_dt is None:
            return text, original_number_format
        if source_type == "datetime":
            return parsed_dt, original_number_format
        return parsed_dt.date(), original_number_format

    if source_type == "boolean":
        lowered = trimmed.lower()
        if lowered in {"true", "yes", "1"}:
            return True, "General"
        if lowered in {"false", "no", "0"}:
            return False, "General"

    return text, original_number_format


def _cleanup_old_tokens_loop():
    """Background loop to delete old token directories by mtime."""
    interval_seconds = 30 * 60  # every 30 minutes
    ttl = timedelta(hours=UPLOAD_TTL_HOURS)
    while True:
        try:
            now = datetime.utcnow()
            if UPLOAD_ROOT.exists():
                for p in UPLOAD_ROOT.iterdir():
                    try:
                        if not p.is_dir():
                            continue
                        mtime = datetime.utcfromtimestamp(p.stat().st_mtime)
                        if now - mtime > ttl:
                            for sub in p.glob("**/*"):
                                try:
                                    if sub.is_file():
                                        sub.unlink(missing_ok=True)
                                except Exception:
                                    pass
                            try:
                                p.rmdir()
                            except Exception:
                                pass
                    except Exception:
                        continue
        except Exception as e:
            logger.warning("cleanup loop error: %s", e)
        time.sleep(interval_seconds)


# Start cleanup thread (daemon)
try:
    t = threading.Thread(target=_cleanup_old_tokens_loop, name="uploads-cleaner", daemon=True)
    t.start()
except Exception as e:
    logger.warning("failed starting cleanup thread: %s", e)


# -----------------
# Simple i18n layer
# -----------------
I18N_DIR = Path(__file__).parent / "i18n"
SUPPORTED_LANGS = ("en", "vi")
_locale_cache: dict[str, dict] = {}


def _load_locale_from_disk(code: str) -> dict:
    try:
        with open(I18N_DIR / f"{code}.json", "r", encoding="utf-8-sig") as f:
            return json.load(f)
    except Exception as exc:
        logger.warning("failed to load locale %s: %s", code, exc)
        return {}


def _get_locale(code: str) -> dict:
    if code not in SUPPORTED_LANGS:
        return {}
    path = I18N_DIR / f"{code}.json"
    try:
        mtime = path.stat().st_mtime
    except FileNotFoundError:
        mtime = None
    cached = _locale_cache.get(code)
    if cached and cached.get("mtime") == mtime:
        return cached.get("data", {})
    data = _load_locale_from_disk(code) if mtime is not None else {}
    _locale_cache[code] = {"mtime": mtime, "data": data}
    return data


def _detect_lang_from_header() -> str:
    header = request.headers.get("Accept-Language", "").lower()
    if header.startswith("vi"):
        return "vi"
    return "en"


def get_lang() -> str:
    lang = request.args.get("lang") or session.get("lang")
    if lang not in SUPPORTED_LANGS:
        lang = _detect_lang_from_header()
    if lang not in SUPPORTED_LANGS:
        lang = "en"
    session["lang"] = lang
    return lang


def tr(key: str) -> str:
    lang = getattr(g, "lang", None) or session.get("lang") or "en"
    locale = _get_locale(lang)
    if key in locale:
        return locale[key]
    fallback = _get_locale("en")
    return fallback.get(key, key)


@app.before_request
def _bind_lang():
    g.lang = get_lang()


@app.context_processor
def _inject_t():
    return {"t": tr, "current_lang": getattr(g, "lang", "en"), "csrf_token": get_csrf_token}


@app.route("/set-lang/<lang>")
def set_lang_route(lang: str):
    if lang not in SUPPORTED_LANGS:
        lang = "en"
    session["lang"] = lang
    nxt = request.args.get("next") or request.referrer or url_for("index")
    try:
        p = urlparse(nxt)
        # Only allow relative paths
        path = p.path or "/"
        unsafe = {url_for("render_view"), url_for("render_multi"), url_for("upload_files")}
        if path in unsafe:
            nxt = url_for("index")
    except Exception:
        nxt = url_for("index")
    return redirect(nxt)


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload_files():
    files = request.files.getlist("files")
    if not files:
        flash(tr("flash_choose_at_least_one_file"))
        return redirect(url_for("index"))

    token = uuid.uuid4().hex
    dest_dir = get_token_dir(token)

    saved_any = False
    for f in files:
        if not f or f.filename == "":
            continue
        if not allowed_file(f.filename):
            continue
        filename = secure_filename(f.filename)
        path = dest_dir / filename
        f.save(path)
        saved_any = True

    if not saved_any:
        flash(tr("flash_no_valid_files"))
        return redirect(url_for("index"))

    return redirect(url_for("select", token=token))


@app.route("/select/<token>", methods=["GET"])
def select(token: str):
    dest_dir = get_token_dir(token)
    files = []
    for p in sorted(dest_dir.iterdir()):
        if not p.is_file():
            continue
        filename = p.name
        if not allowed_file(filename):
            continue
        ext = filename.rsplit(".", 1)[1].lower()
        if ext == "csv":
            sheets = ["CSV"]
            try:
                _, csv_meta = read_csv_smart(p)
            except Exception:
                csv_meta = {}
        else:
            csv_meta = {}
            try:
                xls = pd.ExcelFile(p)
                sheets = list(xls.sheet_names)
            except Exception:
                sheets = []
        files.append({"filename": filename, "ext": ext, "sheets": sheets, "csv_meta": csv_meta})

    if not files:
        flash(tr("flash_uploaded_unreadable"))
        return redirect(url_for("index"))

    return render_template("select.html", token=token, files=files)


@app.route("/render", methods=["POST"])
def render_view():
    token = request.form.get("token")
    selection = request.form.get("selection")
    if not token or not selection:
        flash(tr("flash_invalid_selection"))
        return redirect(url_for("index"))

    if "::" in selection:
        filename, sheet_name = selection.split("::", 1)
    else:
        filename, sheet_name = selection, ""

    dest_dir = get_token_dir(token)
    path = dest_dir / filename
    if not path.exists():
        flash(tr("flash_selected_file_not_found"))
        return redirect(url_for("index"))

    try:
        df, _, target_sheet = _load_sheet_dataframe(path, filename, sheet_name or None)
        table_html = df.to_html(classes="table table-striped table-sm", index=False, border=0, na_rep="")
        return render_template("view.html", filename=filename, sheet_name=target_sheet, table_html=table_html, token=token)
    except Exception as e:
        flash(f"{tr('flash_failed_open_file')}: {e}")
        return redirect(url_for("select", token=token))


@app.route("/render_multi", methods=["POST"])
def render_multi():
    token = request.form.get("token")
    selections = request.form.getlist("selection")
    if not token or not selections:
        flash(tr("flash_select_at_least_one_sheet"))
        return redirect(url_for("index"))

    dest_dir = get_token_dir(token)

    views = []
    for sel in selections:
        if "::" in sel:
            filename, sheet_name = sel.split("::", 1)
        else:
            filename, sheet_name = sel, ""

        path = dest_dir / filename
        if not path.exists():
            continue

        try:
            df, column_metadata, target_sheet = _load_sheet_dataframe(path, filename, sheet_name or None)
            label = f"{filename}" if not target_sheet else f"{filename} - {target_sheet}"
            table_html = df.to_html(classes="table table-striped table-sm", index=False, border=0, na_rep="")
            views.append({
                "id": f"v_{len(views)}",
                "label": label,
                "filename": filename,
                "sheet_name": target_sheet,
                "table_html": table_html,
                "column_metadata": column_metadata,
            })
        except Exception:
            continue

    if not views:
        flash(tr("flash_selected_sheets_could_not_be_opened"))
        return redirect(url_for("select", token=token))

    return render_template("multi_view.html", token=token, views=views, column_format_presets=COLUMN_FORMAT_PRESETS)


def _sanitize_sheet_name(name: str) -> str:
    # Remove invalid characters: : \ / ? * [ ]
    name = re.sub(r"[:\\/\?\*\[\]]", " ", name)
    name = name.strip() or "Sheet"
    return name[:31]


@app.route("/export", methods=["POST"])
def export_excel():
    payload = request.get_json(silent=True) or {}
    sheets = payload.get("sheets")
    out_name = payload.get("filename") or "export.xlsx"
    if not sheets or not isinstance(sheets, list):
        return {"error": tr("flash_select_at_least_one_sheet")}, 400

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used_names = set()

    for item in sheets:
        headers = item.get("headers") or []
        rows = item.get("rows") or []
        column_formats = item.get("column_formats") or []
        sheet_name = item.get("name") or "Sheet"
        sheet_name = _sanitize_sheet_name(sheet_name)
        base = sheet_name
        i = 1
        while sheet_name in used_names:
            suffix = f"_{i}"
            sheet_name = _sanitize_sheet_name(base[: (31 - len(suffix))] + suffix)
            i += 1
        used_names.add(sheet_name)

        worksheet = workbook.create_sheet(title=sheet_name)
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col_idx, value=header)

        for row_idx, row_values in enumerate(rows, start=2):
            for col_idx, raw_value in enumerate(row_values, start=1):
                column_meta = column_formats[col_idx - 1] if col_idx - 1 < len(column_formats) else {}
                value, number_format = _coerce_export_cell(raw_value, column_meta)
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                if number_format:
                    cell.number_format = number_format

        for col_idx, _ in enumerate(headers, start=1):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=out_name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000)






