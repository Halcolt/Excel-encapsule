from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from projects.excel.app import main


def _create_sample_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "People"
    sheet.append(["Name", "Birthday", "Score"])
    sheet.append(["Alice", "2024-03-15", 10])
    sheet.append(["Bob", "2024-11-05", 20])

    extra = workbook.create_sheet("Summary")
    extra.append(["Label", "Value"])
    extra.append(["Count", 2])

    buf = BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _set_test_upload_root(tmp_path: Path) -> None:
    upload_root = tmp_path / "uploads"
    upload_root.mkdir(parents=True, exist_ok=True)
    main.UPLOAD_ROOT = upload_root


def _get_csrf_token(client) -> str:
    client.get("/")
    with client.session_transaction() as session:
        return session["_csrf_token"]


def test_sanitize_sheet_name_trims_invalid_chars():
    assert main._sanitize_sheet_name("  bad:/name*[]  ") == "bad  name"


def test_load_sheet_dataframe_builds_excel_metadata(tmp_path):
    workbook_path = tmp_path / "sample.xlsx"
    workbook_path.write_bytes(_create_sample_workbook_bytes())

    df, metadata, sheet_name = main._load_sheet_dataframe(workbook_path, workbook_path.name, "People")

    assert sheet_name == "People"
    assert list(df.columns) == ["Name", "Birthday", "Score"]
    assert metadata[1]["source_type"] == "text"
    assert metadata[2]["source_type"] == "integer"


def test_upload_select_render_multi_and_export_flow(tmp_path):
    _set_test_upload_root(tmp_path)
    main.app.config["TESTING"] = True
    main.app.secret_key = "test-secret"

    workbook_bytes = _create_sample_workbook_bytes()

    with main.app.test_client() as client:
        csrf_token = _get_csrf_token(client)
        upload_response = client.post(
            "/upload",
            data={
                "csrf_token": csrf_token,
                "files": (BytesIO(workbook_bytes), "sample.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert upload_response.status_code == 302
        assert "/select/" in upload_response.location

        token = upload_response.location.rsplit("/", 1)[-1]
        select_response = client.get(upload_response.location)
        assert select_response.status_code == 200
        assert b"sample.xlsx" in select_response.data

        csrf_token = _get_csrf_token(client)
        render_response = client.post(
            "/render_multi",
            data={
                "csrf_token": csrf_token,
                "token": token,
                "selection": ["sample.xlsx::People", "sample.xlsx::Summary"],
            },
        )
        assert render_response.status_code == 200
        assert b"sample.xlsx - People" in render_response.data
        assert b"sample.xlsx - Summary" in render_response.data

        csrf_token = _get_csrf_token(client)
        export_response = client.post(
            "/export",
            headers={"X-CSRFToken": csrf_token},
            json={
                "filename": "export.xlsx",
                "csrf_token": csrf_token,
                "sheets": [
                    {
                        "headers": ["Name", "Birthday", "Score"],
                        "rows": [
                            ["Alice", "2024-03-15", "10"],
                            ["Bob", "2024-11-05", "20"],
                        ],
                        "name": "People",
                        "column_formats": [
                            {
                                "header": "Name",
                                "source_type": "text",
                                "original_number_format": "@",
                                "selected_preset": "original",
                            },
                            {
                                "header": "Birthday",
                                "source_type": "text",
                                "original_number_format": "@",
                                "selected_preset": "original",
                            },
                            {
                                "header": "Score",
                                "source_type": "integer",
                                "original_number_format": "0",
                                "selected_preset": "original",
                            },
                        ],
                    }
                ],
            },
        )
        assert export_response.status_code == 200

        output = tmp_path / "export.xlsx"
        output.write_bytes(export_response.data)
        workbook = load_workbook(output)
        try:
            sheet = workbook["People"]
            assert sheet["A2"].value == "Alice"
            assert sheet["C2"].value == 10
            assert sheet["C2"].number_format == "0"
        finally:
            workbook.close()
